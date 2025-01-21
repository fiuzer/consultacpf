'''
1 - Entrar na planilha e extrair o cpf do cliente
2 - Entrar no site para usar o cpf para consulta do status do cliente
3 - Verificar se está "em dia" ou "Atrasado"
4 - se estiver em dia, pegar a data de pagamento e o método de pagamento
5 -  Caso contrário, colocar o status como pendente
6 - inserir essas informações em uma nova planilha(nome, valor, cpf, vencimento e status e caso esteja em dia, data pagamento  e método pagamento)
7 - repetir até acabar os clientes.
'''

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# Ler a planilha
planilha_cliente = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_cliente = planilha_cliente['Sheet1']
# Abre o navegador e entra no site: https://consultcpf-devaprender.netlify.app/
browser = webdriver.Chrome()
browser.get('https://consultcpf-devaprender.netlify.app/')
 
 # 2 - Usar o cpf para consulta do status do cliente
for linha in pagina_cliente.iter_rows(min_row=2, values_only=True):
  nome, valor, cpf, vencimento = linha
  sleep(5)

  cpf_element = browser.find_element(By.ID, 'cpfInput')
  sleep(1)
  cpf_element.clear()
  cpf_element.send_keys(cpf)
  #3 - Verificar se está "em dia" ou "Atrasado"
  botao_pesquisar =browser.find_element(By.XPATH, '//button[@class="btn btn-custom btn-lg btn-block mt-3"]')
  sleep(1)
  botao_pesquisar.click()
  sleep(5)
  
  status = browser.find_element(By.XPATH, "//span[@id='statusLabel']")
  #4 - se estiver em dia, pegar a data de pagamento e o método de pagamento
  if status.text == 'em dia':
    data_pagamento = browser.find_element(By.ID, 'paymentDate')
    metodo_pagamento = browser.find_element(By.ID, 'paymentMethod')

    planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
    pagina_fechamento = planilha_fechamento['Sheet1']

    pagina_fechamento.append([nome, valor, cpf,vencimento, 'em dia', data_pagamento.text.split()[3], metodo_pagamento.text.split()[3]])
    planilha_fechamento.save('planilha fechamento.xlsx')

  #5 -  Caso contrário, colocar o status como pendente
  else:
    planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
    pagina_fechamento = planilha_fechamento['Sheet1']

    pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
    planilha_fechamento.save('planilha fechamento.xlsx')


  #6 - inserir essas informações em uma nova planilha(nome, valor, cpf, vencimento e status e caso esteja em dia, data pagamento  e método pagamento)